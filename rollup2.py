#!/anaconda3/bin/python

# https://automatetheboringstuff.com/chapter12/

# import load_workbook
from openpyxl import load_workbook

# set file path
filepath="/Users/jeffreywiedemann/Desktop/Resource_Planning/jeff_for_python.xlsx"

wb=load_workbook(filepath)

sheet = wb['Summary']

A2=sheet['A2']
A1=sheet['A1']

print(A2.value, end=",")
print(A1.value)
